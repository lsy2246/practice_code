# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html


# useful for handling different item types with a single interface
from itemadapter import ItemAdapter
from openpyxl import Workbook


class TianqiPipeline:
    def process_item(self, item, spider):
        return item


class ExcelPipeline:
    def __init__(self):
        self.workbook = Workbook()
        self.worksheets = {}

    def open_spider(self, spider):
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)

    def close_spider(self, spider):
        self.workbook.save('weather.xlsx')
        self.workbook.close()

    def process_item(self, item, spider):
        region = item['region']
        if region not in self.worksheets:
            ws = self.workbook.create_sheet(title=region)
            # 定义列标题的顺序
            ws.append(['省份', '日期时间', '城市', '白天温度', '白天天气', '白天风向', '白天风力',
                       '夜间温度', '夜间天气', '夜间风向', '夜间风力'])
            self.worksheets[region] = ws
        ws = self.worksheets[region]
        # 根据标题顺序将数据添加到对应列
        ws.append([
            item['provincial'], item['data_time'], item['city'],
            item['temperature_1'], item['weather_1'], item['wind_direction_1'], item['wind_power_1'],
            item['temperature_2'], item['weather_2'], item['wind_direction_2'], item['wind_power_2']
        ])
        return item
